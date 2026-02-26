import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from django.db import transaction
from members.models import Member

class ExcelHandler:
    """Handle Excel import / export operations for members"""

    @staticmethod
    def generate_template():
        """Generate excel template for member import"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Member Import Template"

        # Define headers 
        headers = [
            'date',
            'member_number',
            'member_name',
            'phone',
            'dob_bs',
            'citizenship_no',
            'email',
            'profession',
            'facebook_detail',
            'whatsapp_detail',
            'father_name',
            'grandfather_name',
            'spouse_name',
            'spouse_phone',
            'address',
            'ward_no',           
            'business_name',
            'business_address',
            'job_name',
            'job_address'
        ]

        # Style for header row
        header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Write header with style
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Set column width
        ws.column_dimensions[cell.column_letter].width = 20

        # Format member_number column  as TEXT to preserve leading zeros
        if  header == 'member_number':
            for row_num in range(1, 10001):
                ws.cell(row=row_num, column=col_num).number_format = '@'
        

        # Add sample data row
        sample_data = [
            '2024-01-15',           # date (YYYY-MM-DD)
            '001000001',            # member_number (with leading zeros)
            'राम बहादुर श्रेष्ठ',      # member_name
            '9841234567',           # phone
            '2040-05-15',          # dob_bs (BS date)
            '12-01-75-12345',      # citizenship_no
            'ram@example.com',     # email
            'व्यवसायी',              # profession
            'ram.shrestha',        # facebook_detail
            '9841234567',          # whatsapp_detail
            'सुर्य बहादुर श्रेष्ठ',    # father_name
            'धन बहादुर श्रेष्ठ',      # grandfather_name
            'सीता श्रेष्ठ',           # spouse_name
            '9841234568',          # spouse_phone
            'काठमाडौं',             # address
            '5',                   # ward_no
            'श्रेष्ठ इन्टरप्राइजेज',   # business_name
            'काठमाडौं',             # business_address
            'ABC Company',         # job_name
            'काठमाडौं'              # job_address
        ]

        for col_num, value in enumerate(sample_data, 1):
            cell = ws.cell(row=2, column=col_num)
            cell.value = value
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

        # Add instruction sheet
        ws_instructions = wb.create_sheet("Instructions")
        instructions = [
            ["Excel Import Instructions", ""],
            ["", ""],
            ["Required Fields:", ""],
            ["1. date", "Format: YYYY-MM-DD (e.g., 2024-01-15)"],
            ["2. member_number", "Unique member number (e.g., 000000001)"],
            ["3. member_name", "Member's full name"],
            ["", ""],
            ["Optional Fields:", ""],
            ["- phone", "Contact number"],
            ["- dob_bs", "Date of birth in BS (e.g., 2040-05-15)"],
            ["- citizenship_no", "Citizenship number"],
            ["- email", "Email address"],
            ["- All other fields are optional", ""],
            ["", ""],
            ["Important Notes:", ""],
            ["1. Do not modify header row", ""],
            ["2. Delete sample data before importing", ""],
            ["3. Date format must be YYYY-MM-DD", ""],
            ["4. Member numbers must be unique", ""],
            ["5. Save file as .xlsx format", ""],
        ]

        for row_num, (instruction, detail) in enumerate(instructions, 1):
            ws_instructions.cell(row=row_num, column=1).value = instruction
            ws_instructions.cell(row=row_num, column=2).value = detail

            if row_num == 1:
                ws_instructions.cell(row=row_num, column=1).font = Font(bold=True, size=14)
            
        ws_instructions.column_dimensions['A'].width = 30
        ws_instructions.column_dimensions['B'].width = 50

        return wb
    
    @staticmethod
    def validate_excel_data(file_path):
        """Validate Excel file before import"""
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            errors = []
            warnings = []

            # Check if file is empty
            if ws.max_row < 2:
                errors.append("Excel file is empty. No data to import.")
                return False, errors, warnings
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]

            # Required headers
            required_headers = ['date', 'member_number', 'member_name']

            # Check required headers
            for req_header in required_headers:
                if req_header not in headers:
                    errors.append(f"Missing required column: {req_header}")

            if errors:
                return False, errors, warnings
            
            # Validate data rows
            member_numbers = set()
            for row_num in range(2, ws.max_row + 1):
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value
                
                # Check required fields
                if not row_data.get('member_number'):
                    errors.append(f"Row {row_num}: member_number is required!")
                elif row_data['member_number'] in member_numbers:
                    errors.append(f"Row {row_num}: Duplicate member_number {row_data['member_number']}")
                else:
                    member_numbers.add(row_data['member_number'])
                
                if not row_data.get('member_name'):
                    errors.append(f"Row {row_num}: member_name is required!")
                
                if not row_data.get('date'):
                    warnings.append(f"Row {row_num}: date is missing,, will use today's date")
                
                # Validate date format
                if row_data.get('date'):
                    try:
                        if isinstance(row_data['date'], str):
                            datetime.strptime(row_data['date'], '%Y-%m-%d')
                    except ValueError:
                        errors.append(f"Row {row_num}: Invalid date format. Use YYYY-MM_DD")
            
            if errors:
                return False, errors, warnings
            
            return True, [], warnings
            
        except Exception as e:
            return False, [f"Error reading Excel file: {str(e)}"], []
        

    @staticmethod
    @transaction.atomic
    def import_from_excel(file_path):
        """Import members from Excel file"""
        try:
            # Validate first
            is_valid, errors, warnings = ExcelHandler.validate_excel_data(file_path)
            if not is_valid:
                return False, 'Data validation failed!', errors, warnings
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            # Get headers
            headers = [cell.value for cell in ws[1]]

            created_count = 0
            updated_count = 0
            skipped_count = 0

            # Process each row
            for row_num in range(2, ws.max_row + 1):
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_data[header] = cell_value if cell_value else ''
                
                # Skip empty rows
                if not row_data.get('member_number'):
                    skipped_count += 1
                    continue

                # Prepare member data
                member_number = str(row_data['member_number']).strip()

                # Handle leading zeros: pad to 9 digits if numeric
                if member_number.isdigit():
                    member_number = member_number.zfill(9)

                # Parse Date
                date_value = row_data.get('date')    
                if isinstance(date_value, str):
                    try:
                        date_value = datetime.strptime(date_value, '%Y-%m-%d').date()
                    except ValueError:
                        date_value = datetime.now().date()
                elif not date_value:
                    date_value = datetime.now().date()
                
                member_data = {
                    'date': date_value,
                    'member_name': row_data.get('member_name', ''),
                    'phone': row_data.get('phone', ''),
                    'dob_bs': row_data.get('dob_bs', ''),
                    'citizenship_no': row_data.get('citizenship_no', ''),
                    'email': row_data.get('email', ''),
                    'profession': row_data.get('profession', ''),
                    'facebook_detail': row_data.get('facebook_detail', ''),
                    'whatsapp_detail': row_data.get('whatsapp_detail', ''),
                    'father_name': row_data.get('father_name', ''),
                    'grandfather_name': row_data.get('grandfather_name', ''),
                    'spouse_name': row_data.get('spouse_name', ''),
                    'spouse_phone': row_data.get('spouse_phone', ''),
                    'address': row_data.get('address', ''),
                    'ward_no': row_data.get('ward_no', ''),
                    'business_name': row_data.get('business_name', ''),
                    'business_address': row_data.get('business_address', ''),
                    'job': row_data.get('job_name', ''),
                    'job_address': row_data.get('job_address', ''),

                }

                # Check if member exists
                try:
                    member = Member.objects.get(member_number=member_number)
                    # Update existing member
                    for key, value in member_data.items():
                        setattr(member, key, value)
                    member.save()
                    updated_count += 1
                except Member.DoesNotExist:
                    # Create new member
                    Member.objects.create(
                        member_number=member_number,
                        **member_data
                    )
                    created_count += 1
            message = f"Import successful! Created: {created_count}, Updated: {updated_count}, Skipped: {skipped_count}" 
            return True, message, [], warnings
        
        except Exception as e:
            return False, f"Import failed: {str(e)}", [], []